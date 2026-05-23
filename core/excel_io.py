"""Acceso de bajo nivel a la planilla: cargar, guardar con backup, buscar fila libre."""
import os, shutil
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", str(ROOT / "Finanzas_Familia_2026.xlsx")))
BACKUP_DIR = ROOT / "backups"
HDR = 4  # encabezados en fila 4; datos desde 5

def parse_fecha(s):
    if not s:
        return date.today()
    if isinstance(s, (datetime, date)):
        return s if isinstance(s, date) else s.date()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(s), fmt).date()
        except ValueError:
            continue
    return date.today()

def tipo_cambio():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        return float(wb["Parametros"]["B4"].value or 0)
    finally:
        wb.close()

def inquilinos():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    out = []
    for row in wb["Inquilinos"].iter_rows(min_row=HDR + 1, max_row=HDR + 29):
        idv, nombre, local = row[0].value, row[1].value, row[2].value
        if idv:
            out.append({"id": idv, "nombre": nombre, "local": local})
    wb.close()
    return out

def resolver_id(id_inquilino=0, texto=""):
    if id_inquilino:
        return int(id_inquilino)
    t = (texto or "").strip().lower()
    if not t:
        return None
    for inq in inquilinos():
        if t == str(inq["id"]) or t in str(inq["nombre"]).lower() or t in str(inq["local"]).lower() \
           or str(inq["nombre"]).lower() in t or str(inq["local"]).lower() in t:
            return inq["id"]
    return None

def backup():
    BACKUP_DIR.mkdir(exist_ok=True)
    dst = BACKUP_DIR / f"{EXCEL_PATH.stem}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(EXCEL_PATH, dst)
    return dst

def _next_row(ws, key_col, start, end):
    r = start
    while r <= end and ws.cell(row=r, column=key_col).value not in (None, ""):
        r += 1
    return r

def open_write():
    try:
        return load_workbook(EXCEL_PATH)
    except PermissionError:
        raise RuntimeError("La planilla está abierta en Excel. Cerrala y reintentá.")

def save(wb):
    wb.calculation.fullCalcOnLoad = True
    wb.save(EXCEL_PATH)
