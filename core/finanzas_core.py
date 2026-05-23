"""Las TOOLS del sistema: funciones que leen/escriben la planilla de finanzas.

Las usan tanto el servidor MCP (para Claude) como la app web (para Gemini).
Cada función está documentada con type hints para que el LLM arme sus declaraciones.
"""
from datetime import date
from collections import defaultdict
from openpyxl import load_workbook
from . import excel_io as io

_LOG = []  # registro de operaciones de la última corrida

def get_and_clear_log():
    global _LOG
    out, _LOG = _LOG, []
    return out

def listar_inquilinos() -> list:
    """Devuelve la lista de inquilinos (id, nombre, local) para identificar cobros."""
    return io.inquilinos()

def agregar_cobro(monto: float, inquilino: str = "", id_inquilino: int = 0,
                  fecha: str = "", moneda: str = "ARS", medio_pago: str = "",
                  estado: str = "Cobrado", periodo: str = "", notas: str = "") -> dict:
    """Registra el cobro del alquiler de un inquilino en la hoja Cobros.

    Args:
        monto: importe cobrado.
        inquilino: nombre o local del inquilino (si no sabés el id).
        id_inquilino: id numérico (1-29) si lo conocés.
        fecha: fecha del cobro en formato YYYY-MM-DD (si falta, hoy).
        moneda: "ARS" o "USD".
        medio_pago: efectivo, transferencia, cheque, etc.
        estado: "Cobrado", "Pendiente", "Parcial" o "Vencido".
        periodo: período al que corresponde (texto, ej "Mayo 2026").
        notas: observaciones.
    """
    idn = io.resolver_id(id_inquilino, inquilino)
    wb = io.open_write()
    ws = wb["Cobros"]
    r = io._next_row(ws, 1, io.HDR + 1, io.HDR + 400)
    ws.cell(row=r, column=1, value=io.parse_fecha(fecha))
    ws.cell(row=r, column=2, value=idn)
    ws.cell(row=r, column=5, value=periodo or None)
    ws.cell(row=r, column=6, value=monto)
    ws.cell(row=r, column=7, value=(moneda or "ARS").upper())
    ws.cell(row=r, column=10, value=medio_pago or None)
    ws.cell(row=r, column=11, value=estado or "Cobrado")
    ws.cell(row=r, column=12, value=notas or None)
    io.backup(); io.save(wb)
    res = {"hoja": "Cobros", "fila": r, "id_inquilino": idn, "monto": monto, "moneda": moneda}
    _LOG.append(res)
    return res

def agregar_movimiento(monto: float, tipo: str = "Egreso", miembro: str = "Familia",
                       categoria: str = "", descripcion: str = "", fecha: str = "",
                       moneda: str = "ARS", medio_pago: str = "", notas: str = "") -> dict:
    """Registra un ingreso o egreso de la familia en la hoja Movimientos.

    Args:
        monto: importe.
        tipo: "Ingreso" o "Egreso".
        miembro: "Yo", "Hermana", "Mamá" o "Familia".
        categoria: categoría del gasto/ingreso (ej "Supermercado/Comida").
        descripcion: detalle.
        fecha: YYYY-MM-DD (si falta, hoy).
        moneda: "ARS" o "USD".
        medio_pago: medio de pago.
        notas: observaciones.
    """
    wb = io.open_write()
    ws = wb["Movimientos"]
    r = io._next_row(ws, 1, io.HDR + 1, io.HDR + 400)
    ws.cell(row=r, column=1, value=io.parse_fecha(fecha))
    ws.cell(row=r, column=3, value=tipo or "Egreso")
    ws.cell(row=r, column=4, value=miembro or "Familia")
    ws.cell(row=r, column=5, value=categoria or None)
    ws.cell(row=r, column=6, value=descripcion or None)
    ws.cell(row=r, column=7, value=monto)
    ws.cell(row=r, column=8, value=(moneda or "ARS").upper())
    ws.cell(row=r, column=10, value=medio_pago or None)
    ws.cell(row=r, column=11, value=notas or None)
    io.backup(); io.save(wb)
    res = {"hoja": "Movimientos", "fila": r, "tipo": tipo, "monto": monto, "categoria": categoria}
    _LOG.append(res)
    return res

def actualizar_ipc(mes: str, indice: float) -> dict:
    """Carga el índice IPC del INDEC para un mes en la hoja Inflacion INDEC.

    Args:
        mes: mes en formato YYYY-MM (ej "2026-04").
        indice: valor del índice IPC nivel general publicado por INDEC.
    """
    from datetime import datetime as _dt
    wb = io.open_write()
    ws = wb["Inflacion INDEC"]
    mm = (mes or "")[:7]
    target = None
    for row in ws.iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 60):
        b = row[1].value
        if hasattr(b, "strftime") and b.strftime("%Y-%m") == mm:
            target = row[1].row
            break
    if not target:
        raise RuntimeError(f"El mes {mm} no está en la tabla de Inflacion INDEC (rango 2022-2026).")
    ws.cell(row=target, column=3, value=indice)
    io.backup(); io.save(wb)
    res = {"hoja": "Inflacion INDEC", "fila": target, "mes": mm, "indice": indice}
    _LOG.append(res)
    return res

def resumen_mensual(mes: int = 0) -> dict:
    """Resumen financiero calculado en Python desde los datos cargados.

    Args:
        mes: número de mes 1-12 para filtrar; 0 = todo el año.
    Devuelve ingresos por alquileres, otros ingresos, egresos y resultado neto (ARS).
    """
    tc = io.tipo_cambio() or 1
    wb = load_workbook(io.EXCEL_PATH, read_only=True, data_only=True)
    alq = 0.0
    for row in wb["Cobros"].iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 400):
        f, monto, moneda = row[0].value, row[5].value, row[6].value
        if monto and hasattr(f, "month") and (mes == 0 or f.month == mes):
            alq += monto * (tc if str(moneda).upper() == "USD" else 1)
    ing = egr = 0.0
    for row in wb["Movimientos"].iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 400):
        f, tipo, monto, moneda = row[0].value, row[2].value, row[6].value, row[7].value
        if monto and hasattr(f, "month") and (mes == 0 or f.month == mes):
            ars = monto * (tc if str(moneda).upper() == "USD" else 1)
            if str(tipo).lower() == "ingreso":
                ing += ars
            else:
                egr += ars
    wb.close()
    total_ing = alq + ing
    return {
        "mes": mes or "año completo",
        "ingresos_alquileres": round(alq),
        "otros_ingresos": round(ing),
        "total_ingresos": round(total_ing),
        "egresos": round(egr),
        "resultado_neto": round(total_ing - egr),
        "moneda": "ARS",
    }

# lista de tools para registrar en Gemini / MCP
TOOLS = [agregar_cobro, agregar_movimiento, actualizar_ipc, listar_inquilinos, resumen_mensual]
