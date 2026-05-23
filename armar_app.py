"""Scaffolder del proyecto Finanzas Familia (Gemini + tools + MCP + núcleo compartido).

Genera/re-genera: core/ (las tools y el acceso al Excel), mcp/ (servidor MCP para
usar desde Claude), app/ (web con Gemini para audio/texto/documentos), scripts y docs.
Es idempotente. Tambien es el "deliverable #1": el script que arma/modifica el proyecto.
"""
from pathlib import Path

BASE = Path(__file__).resolve().parent
FILES = {}

# ============================================================ core
FILES["core/__init__.py"] = "\n"

FILES["core/excel_io.py"] = r'''"""Acceso de bajo nivel a la planilla: cargar, guardar con backup, buscar fila libre."""
import os, shutil
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parents[1] / ".env")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", str(ROOT / "Finanzas_Familia_2026.xlsx")))
BACKUP_DIR = ROOT / "backups"
HDR = 4  # encabezados en fila 4; datos desde 5

def parse_fecha(s):
    if not s:
        return date.today()
    if isinstance(s, (datetime, date)):
        return s if isinstance(s, date) else s.date()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(s), fmt).date()
        except ValueError:
            continue
    return date.today()

def tipo_cambio():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        return float(wb["Parametros"]["B4"].value or 0)
    finally:
        wb.close()

def inquilinos():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    out = []
    for row in wb["Inquilinos"].iter_rows(min_row=HDR + 1, max_row=HDR + 29):
        idv, nombre, local = row[0].value, row[1].value, row[2].value
        if idv:
            out.append({"id": idv, "nombre": nombre, "local": local})
    wb.close()
    return out

def resolver_id(id_inquilino=0, texto=""):
    if id_inquilino:
        return int(id_inquilino)
    t = (texto or "").strip().lower()
    if not t:
        return None
    for inq in inquilinos():
        if t == str(inq["id"]) or t in str(inq["nombre"]).lower() or t in str(inq["local"]).lower() \
           or str(inq["nombre"]).lower() in t or str(inq["local"]).lower() in t:
            return inq["id"]
    return None

def backup():
    BACKUP_DIR.mkdir(exist_ok=True)
    dst = BACKUP_DIR / f"{EXCEL_PATH.stem}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(EXCEL_PATH, dst)
    return dst

def _next_row(ws, key_col, start, end):
    r = start
    while r <= end and ws.cell(row=r, column=key_col).value not in (None, ""):
        r += 1
    return r

def open_write():
    try:
        return load_workbook(EXCEL_PATH)
    except PermissionError:
        raise RuntimeError("La planilla está abierta en Excel. Cerrala y reintentá.")

def save(wb):
    wb.calculation.fullCalcOnLoad = True
    wb.save(EXCEL_PATH)
'''

FILES["core/finanzas_core.py"] = r'''"""Las TOOLS del sistema: funciones que leen/escriben la planilla de finanzas.

Las usan tanto el servidor MCP (para Claude) como la app web (para Gemini).
Cada función está documentada con type hints para que el LLM arme sus declaraciones.
"""
from datetime import date
from collections import defaultdict
from openpyxl import load_workbook
from . import excel_io as io

_LOG = []  # registro de operaciones de la última corrida

def get_and_clear_log():
    global _LOG
    out, _LOG = _LOG, []
    return out

def listar_inquilinos() -> list:
    """Devuelve la lista de inquilinos (id, nombre, local) para identificar cobros."""
    return io.inquilinos()

def agregar_cobro(monto: float, inquilino: str = "", id_inquilino: int = 0,
                  fecha: str = "", moneda: str = "ARS", medio_pago: str = "",
                  estado: str = "Cobrado", periodo: str = "", notas: str = "") -> dict:
    """Registra el cobro del alquiler de un inquilino en la hoja Cobros.

    Args:
        monto: importe cobrado.
        inquilino: nombre o local del inquilino (si no sabés el id).
        id_inquilino: id numérico (1-29) si lo conocés.
        fecha: fecha del cobro en formato YYYY-MM-DD (si falta, hoy).
        moneda: "ARS" o "USD".
        medio_pago: efectivo, transferencia, cheque, etc.
        estado: "Cobrado", "Pendiente", "Parcial" o "Vencido".
        periodo: período al que corresponde (texto, ej "Mayo 2026").
        notas: observaciones.
    """
    idn = io.resolver_id(id_inquilino, inquilino)
    wb = io.open_write()
    ws = wb["Cobros"]
    r = io._next_row(ws, 1, io.HDR + 1, io.HDR + 400)
    ws.cell(row=r, column=1, value=io.parse_fecha(fecha))
    ws.cell(row=r, column=2, value=idn)
    ws.cell(row=r, column=5, value=periodo or None)
    ws.cell(row=r, column=6, value=monto)
    ws.cell(row=r, column=7, value=(moneda or "ARS").upper())
    ws.cell(row=r, column=10, value=medio_pago or None)
    ws.cell(row=r, column=11, value=estado or "Cobrado")
    ws.cell(row=r, column=12, value=notas or None)
    io.backup(); io.save(wb)
    res = {"hoja": "Cobros", "fila": r, "id_inquilino": idn, "monto": monto, "moneda": moneda}
    _LOG.append(res)
    return res

def agregar_movimiento(monto: float, tipo: str = "Egreso", miembro: str = "Familia",
                       categoria: str = "", descripcion: str = "", fecha: str = "",
                       moneda: str = "ARS", medio_pago: str = "", notas: str = "") -> dict:
    """Registra un ingreso o egreso de la familia en la hoja Movimientos.

    Args:
        monto: importe.
        tipo: "Ingreso" o "Egreso".
        miembro: "Yo", "Hermana", "Mamá" o "Familia".
        categoria: categoría del gasto/ingreso (ej "Supermercado/Comida").
        descripcion: detalle.
        fecha: YYYY-MM-DD (si falta, hoy).
        moneda: "ARS" o "USD".
        medio_pago: medio de pago.
        notas: observaciones.
    """
    wb = io.open_write()
    ws = wb["Movimientos"]
    r = io._next_row(ws, 1, io.HDR + 1, io.HDR + 400)
    ws.cell(row=r, column=1, value=io.parse_fecha(fecha))
    ws.cell(row=r, column=3, value=tipo or "Egreso")
    ws.cell(row=r, column=4, value=miembro or "Familia")
    ws.cell(row=r, column=5, value=categoria or None)
    ws.cell(row=r, column=6, value=descripcion or None)
    ws.cell(row=r, column=7, value=monto)
    ws.cell(row=r, column=8, value=(moneda or "ARS").upper())
    ws.cell(row=r, column=10, value=medio_pago or None)
    ws.cell(row=r, column=11, value=notas or None)
    io.backup(); io.save(wb)
    res = {"hoja": "Movimientos", "fila": r, "tipo": tipo, "monto": monto, "categoria": categoria}
    _LOG.append(res)
    return res

def actualizar_ipc(mes: str, indice: float) -> dict:
    """Carga el índice IPC del INDEC para un mes en la hoja Inflacion INDEC.

    Args:
        mes: mes en formato YYYY-MM (ej "2026-04").
        indice: valor del índice IPC nivel general publicado por INDEC.
    """
    from datetime import datetime as _dt
    wb = io.open_write()
    ws = wb["Inflacion INDEC"]
    mm = (mes or "")[:7]
    target = None
    for row in ws.iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 60):
        b = row[1].value
        if hasattr(b, "strftime") and b.strftime("%Y-%m") == mm:
            target = row[1].row
            break
    if not target:
        raise RuntimeError(f"El mes {mm} no está en la tabla de Inflacion INDEC (rango 2022-2026).")
    ws.cell(row=target, column=3, value=indice)
    io.backup(); io.save(wb)
    res = {"hoja": "Inflacion INDEC", "fila": target, "mes": mm, "indice": indice}
    _LOG.append(res)
    return res

def resumen_mensual(mes: int = 0) -> dict:
    """Resumen financiero calculado en Python desde los datos cargados.

    Args:
        mes: número de mes 1-12 para filtrar; 0 = todo el año.
    Devuelve ingresos por alquileres, otros ingresos, egresos y resultado neto (ARS).
    """
    tc = io.tipo_cambio() or 1
    wb = load_workbook(io.EXCEL_PATH, read_only=True, data_only=True)
    alq = 0.0
    for row in wb["Cobros"].iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 400):
        f, monto, moneda = row[0].value, row[5].value, row[6].value
        if monto and hasattr(f, "month") and (mes == 0 or f.month == mes):
            alq += monto * (tc if str(moneda).upper() == "USD" else 1)
    ing = egr = 0.0
    for row in wb["Movimientos"].iter_rows(min_row=io.HDR + 1, max_row=io.HDR + 400):
        f, tipo, monto, moneda = row[0].value, row[2].value, row[6].value, row[7].value
        if monto and hasattr(f, "month") and (mes == 0 or f.month == mes):
            ars = monto * (tc if str(moneda).upper() == "USD" else 1)
            if str(tipo).lower() == "ingreso":
                ing += ars
            else:
                egr += ars
    wb.close()
    total_ing = alq + ing
    return {
        "mes": mes or "año completo",
        "ingresos_alquileres": round(alq),
        "otros_ingresos": round(ing),
        "total_ingresos": round(total_ing),
        "egresos": round(egr),
        "resultado_neto": round(total_ing - egr),
        "moneda": "ARS",
    }

# lista de tools para registrar en Gemini / MCP
TOOLS = [agregar_cobro, agregar_movimiento, actualizar_ipc, listar_inquilinos, resumen_mensual]
'''

# ============================================================ MCP
FILES["mcp/requirements.txt"] = "mcp>=1.2\nopenpyxl>=3.1\npython-dotenv>=1.0\n"

FILES["mcp/server.py"] = r'''"""Servidor MCP de Finanzas Familia. Expone las tools del núcleo a Claude.

Conectar en Claude Desktop (claude_desktop_config.json):

  "mcpServers": {
    "finanzas-familia": {
      "command": "py",
      "args": ["C:\\Users\\marti\\documents\\proyectojose\\mcp\\server.py"]
    }
  }

Luego, desde Claude, podés escribir: "Cargá el cobro del Local 12, 350 mil".
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from mcp.server.fastmcp import FastMCP
from core import finanzas_core as core

mcp = FastMCP("finanzas-familia")

@mcp.tool()
def agregar_cobro(monto: float, inquilino: str = "", id_inquilino: int = 0, fecha: str = "",
                  moneda: str = "ARS", medio_pago: str = "", estado: str = "Cobrado",
                  periodo: str = "", notas: str = "") -> dict:
    """Registra el cobro del alquiler de un inquilino (hoja Cobros)."""
    return core.agregar_cobro(monto, inquilino, id_inquilino, fecha, moneda,
                              medio_pago, estado, periodo, notas)

@mcp.tool()
def agregar_movimiento(monto: float, tipo: str = "Egreso", miembro: str = "Familia",
                       categoria: str = "", descripcion: str = "", fecha: str = "",
                       moneda: str = "ARS", medio_pago: str = "", notas: str = "") -> dict:
    """Registra un ingreso o egreso de la familia (hoja Movimientos)."""
    return core.agregar_movimiento(monto, tipo, miembro, categoria, descripcion,
                                   fecha, moneda, medio_pago, notas)

@mcp.tool()
def actualizar_ipc(mes: str, indice: float) -> dict:
    """Carga el índice IPC del INDEC para un mes (YYYY-MM)."""
    return core.actualizar_ipc(mes, indice)

@mcp.tool()
def listar_inquilinos() -> list:
    """Lista los inquilinos (id, nombre, local)."""
    return core.listar_inquilinos()

@mcp.tool()
def resumen_mensual(mes: int = 0) -> dict:
    """Resumen de ingresos, egresos y resultado neto (mes 1-12, o 0 = año)."""
    return core.resumen_mensual(mes)

if __name__ == "__main__":
    mcp.run()
'''

FILES["mcp/README_MCP.md"] = r'''# Servidor MCP — Finanzas Familia

Expone las tools del núcleo (`core/finanzas_core.py`) para usarlas **desde Claude**
escribiéndole en lenguaje natural.

## Instalar
```
pip install -r mcp/requirements.txt
```

## Conectar en Claude Desktop
Editá `claude_desktop_config.json` (Configuración → Developer → Edit Config) y agregá:
```json
{
  "mcpServers": {
    "finanzas-familia": {
      "command": "py",
      "args": ["C:\\Users\\marti\\documents\\proyectojose\\mcp\\server.py"]
    }
  }
}
```
Reiniciá Claude Desktop. Vas a ver las tools disponibles y podés decir:
"Cargá un egreso de 48.500 de luz pagado con débito".

## Usarlo en la nube (claude.ai)
Para claude.ai web hay que **hostear** el server como MCP remoto (HTTP/SSE) en internet
con autenticación. Es un paso aparte; pedímelo cuando quieras y te guío con el deploy.
'''

# ============================================================ APP (Gemini)
FILES["app/backend/requirements.txt"] = r'''fastapi>=0.110
uvicorn[standard]>=0.29
python-multipart>=0.0.9
google-genai>=0.3
openpyxl>=3.1
python-dotenv>=1.0
'''

FILES["app/backend/config.py"] = r'''"""Configuración de la app web."""
import os
from pathlib import Path
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parents[2] / ".env")
except Exception:
    pass

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
'''

FILES["app/backend/gemini_agent.py"] = r'''"""Agente Gemini con function calling automático sobre las tools del núcleo.

Gemini entiende audio, imágenes y PDF de forma nativa, así que la misma llamada
sirve para texto, comprobantes (foto/PDF) o audios.
"""
import sys
from pathlib import Path
from datetime import date
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from google import genai
from google.genai import types
from core import finanzas_core as core
import config

SYSTEM = """Sos el asistente de finanzas de una familia que vive de alquileres de locales.
Hoy es {fecha}. Cuando el usuario te manda un audio, texto o comprobante, identificá las
operaciones y cargalas usando las herramientas disponibles:
- Cobros de alquiler -> agregar_cobro (identificá al inquilino por nombre o local; usá listar_inquilinos si hace falta).
- Gastos o ingresos de la familia -> agregar_movimiento.
- Dato del IPC del INDEC -> actualizar_ipc (mes YYYY-MM).
Reglas: si no dicen el año, asumí el actual; moneda por defecto ARS; para egresos elegí
una categoría razonable. Después de cargar, respondé en español qué registraste."""

def procesar(texto: str = "", file_bytes: bytes = None, mime: str = "", filename: str = "") -> dict:
    if not config.GEMINI_API_KEY:
        raise RuntimeError("Falta GEMINI_API_KEY en .env (sacala de aistudio.google.com).")
    core.get_and_clear_log()
    client = genai.Client(api_key=config.GEMINI_API_KEY)
    parts = []
    if file_bytes:
        parts.append(types.Part.from_bytes(data=file_bytes, mime_type=mime or "application/octet-stream"))
    parts.append(types.Part.from_text(text=texto or "Procesá el comprobante adjunto y cargá lo que corresponda."))
    resp = client.models.generate_content(
        model=config.GEMINI_MODEL,
        contents=[types.Content(role="user", parts=parts)],
        config=types.GenerateContentConfig(
            system_instruction=SYSTEM.format(fecha=date.today()),
            tools=core.TOOLS,
        ),
    )
    return {"resumen": (resp.text or "").strip(), "operaciones": core.get_and_clear_log()}
'''

FILES["app/backend/main.py"] = r'''"""API web: subir audio/texto/documento, procesar con Gemini y descargar la planilla."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from core import excel_io
import gemini_agent

app = FastAPI(title="Finanzas Familia")
FRONT = Path(__file__).resolve().parents[1] / "frontend" / "index.html"

@app.get("/", response_class=HTMLResponse)
def home():
    return FRONT.read_text(encoding="utf-8")

@app.post("/procesar")
async def procesar(texto: str = Form(""), archivo: UploadFile = File(None)):
    try:
        data = mime = name = None
        if archivo is not None and archivo.filename:
            data = await archivo.read()
            mime = archivo.content_type
            name = archivo.filename
        res = gemini_agent.procesar(texto=texto or "", file_bytes=data, mime=mime, filename=name)
        return JSONResponse(res)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

@app.get("/excel")
def descargar_excel():
    return FileResponse(
        excel_io.EXCEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=excel_io.EXCEL_PATH.name,
    )

@app.get("/resumen")
def resumen():
    from core import finanzas_core
    return finanzas_core.resumen_mensual(0)
'''

FILES["app/frontend/index.html"] = r'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Finanzas Familia · Carga rápida</title>
<style>
  :root { --azul:#1F3864; --azul2:#2e5090; --bg:#f4f6fb; }
  * { box-sizing:border-box; font-family:system-ui,Segoe UI,Roboto,sans-serif; }
  body { margin:0; background:var(--bg); color:#1b1b1b; }
  header { background:var(--azul); color:#fff; padding:20px 24px; }
  header h1 { margin:0; font-size:20px; }
  header p { margin:4px 0 0; opacity:.85; font-size:13px; }
  main { max-width:720px; margin:24px auto; padding:0 16px; }
  .card { background:#fff; border-radius:14px; padding:20px; box-shadow:0 2px 10px rgba(0,0,0,.06); margin-bottom:18px; }
  textarea { width:100%; min-height:90px; border:1px solid #cfd6e4; border-radius:10px; padding:12px; font-size:15px; resize:vertical; }
  .row { display:flex; gap:10px; flex-wrap:wrap; margin-top:12px; align-items:center; }
  button { border:0; border-radius:10px; padding:11px 16px; font-size:15px; cursor:pointer; }
  .primary { background:var(--azul); color:#fff; }
  .ghost { background:#eef1f8; color:var(--azul); }
  .rec { background:#c0392b; color:#fff; }
  .rec.on { animation:pulse 1s infinite; }
  @keyframes pulse { 50% { opacity:.55; } }
  #salida { white-space:pre-wrap; font-size:14px; }
  .op { border-left:3px solid var(--azul2); padding:6px 10px; margin:8px 0; background:#f8fafc; border-radius:6px; font-size:13px; }
  .muted{ color:#667; font-size:13px; }
  a.dl { text-decoration:none; }
</style>
</head>
<body>
<header>
  <h1>💰 Finanzas Familia — carga rápida</h1>
  <p>Mandá un audio, un texto o una foto/PDF. Gemini lo entiende y lo carga en la planilla.</p>
</header>
<main>
  <div class="card">
    <textarea id="texto" placeholder="Ej: Cobré el alquiler del Local 12, 350 mil por transferencia. / Pagué la luz 48.500 con débito. / El IPC de abril fue índice 7600."></textarea>
    <div class="row">
      <button class="rec" id="btnRec">🎤 Grabar audio</button>
      <input type="file" id="archivo" accept="audio/*,image/*,.pdf">
    </div>
    <div class="row">
      <button class="primary" id="btnEnviar">Procesar y cargar</button>
      <a class="dl" href="/excel"><button class="ghost" type="button">⬇️ Descargar Excel actual</button></a>
    </div>
    <p class="muted" id="estado"></p>
  </div>
  <div class="card" id="cardSalida" style="display:none">
    <strong>Resultado</strong>
    <div id="salida"></div>
  </div>
</main>
<script>
const $ = id => document.getElementById(id);
let mediaRecorder, chunks = [], audioBlob = null;
$("btnRec").onclick = async () => {
  if (mediaRecorder && mediaRecorder.state === "recording") { mediaRecorder.stop(); return; }
  try {
    const stream = await navigator.mediaDevices.getUserMedia({ audio: true });
    mediaRecorder = new MediaRecorder(stream); chunks = [];
    mediaRecorder.ondataavailable = e => chunks.push(e.data);
    mediaRecorder.onstop = () => {
      audioBlob = new Blob(chunks, { type: "audio/webm" });
      $("estado").textContent = "🎧 Audio grabado, listo para procesar.";
      $("btnRec").textContent = "🎤 Grabar audio"; $("btnRec").classList.remove("on");
      stream.getTracks().forEach(t => t.stop());
    };
    mediaRecorder.start();
    $("btnRec").textContent = "⏹️ Detener"; $("btnRec").classList.add("on");
    $("estado").textContent = "Grabando...";
  } catch (e) { $("estado").textContent = "No pude acceder al micrófono: " + e.message; }
};
$("btnEnviar").onclick = async () => {
  const fd = new FormData();
  fd.append("texto", $("texto").value || "");
  const file = $("archivo").files[0];
  if (file) fd.append("archivo", file);
  else if (audioBlob) fd.append("archivo", audioBlob, "audio.webm");
  $("estado").textContent = "Procesando con Gemini..."; $("btnEnviar").disabled = true;
  try {
    const r = await fetch("/procesar", { method: "POST", body: fd });
    const data = await r.json();
    $("cardSalida").style.display = "block";
    if (data.error) {
      $("salida").innerHTML = "<span style='color:#c0392b'>⚠️ " + data.error + "</span>";
    } else {
      let html = "<p>" + (data.resumen || "") + "</p>";
      (data.operaciones || []).forEach(o => {
        html += "<div class='op'><b>" + o.hoja + "</b> (fila " + o.fila + "): " + JSON.stringify(o) + "</div>";
      });
      if (!(data.operaciones || []).length) html += "<div class='muted'>No se cargó ninguna fila.</div>";
      $("salida").innerHTML = html;
    }
    $("estado").textContent = "Listo."; audioBlob = null; $("texto").value = ""; $("archivo").value = "";
  } catch (e) { $("estado").textContent = "Error: " + e.message; }
  finally { $("btnEnviar").disabled = false; }
};
</script>
</body>
</html>
'''

# ============================================================ scripts / docs
FILES["scripts/modificar_proyecto.py"] = r'''"""CLI de mantenimiento.

  py scripts/modificar_proyecto.py regenerar   # regenera la planilla
  py scripts/modificar_proyecto.py validar      # valida fórmulas
  py scripts/modificar_proyecto.py scaffold      # regenera app/mcp/core
  py scripts/modificar_proyecto.py app           # levanta la app web (Gemini)
  py scripts/modificar_proyecto.py mcp           # corre el servidor MCP (para Claude)
  py scripts/modificar_proyecto.py backup        # copia la planilla a backups/
"""
import subprocess, sys, shutil
from datetime import datetime
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]

def _run(args, cwd=None):
    print(">", " ".join(args)); return subprocess.call(args, cwd=str(cwd or ROOT))

def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "help"
    if cmd == "regenerar": _run(["py", "generar_finanzas.py"])
    elif cmd == "validar": _run(["py", "validar.py"])
    elif cmd == "scaffold": _run(["py", "armar_app.py"])
    elif cmd == "app": _run(["uvicorn", "main:app", "--reload", "--port", "8000"], cwd=ROOT / "app" / "backend")
    elif cmd == "mcp": _run(["py", "mcp/server.py"])
    elif cmd == "backup":
        x = ROOT / "Finanzas_Familia_2026.xlsx"; d = ROOT / "backups"; d.mkdir(exist_ok=True)
        out = d / f"{x.stem}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"; shutil.copy2(x, out); print("Backup:", out)
    else: print(__doc__)

if __name__ == "__main__":
    main()
'''

FILES[".env.example"] = r'''# Copiá como .env y completá
GEMINI_API_KEY=AIza...          # https://aistudio.google.com/apikey
GEMINI_MODEL=gemini-2.5-flash
# EXCEL_PATH=C:\ruta\a\Finanzas_Familia_2026.xlsx
'''

FILES["run.ps1"] = r'''# Levanta la app web (Windows / PowerShell)
$ErrorActionPreference = "Stop"
Set-Location $PSScriptRoot
if (-not (Test-Path ".venv")) { py -m venv .venv }
& ".\.venv\Scripts\python.exe" -m pip install --quiet -r app\backend\requirements.txt
if (-not (Test-Path ".env")) { Copy-Item ".env.example" ".env"; Write-Host "Creé .env: completá GEMINI_API_KEY." -ForegroundColor Yellow }
& ".\.venv\Scripts\python.exe" -m uvicorn main:app --reload --port 8000 --app-dir app\backend
'''

FILES["README.md"] = r'''# Finanzas Familia

Economía doméstica de una familia que vive de alquileres (29 inquilinos), con reajuste
por inflación INDEC y carga de datos por **audio, texto o documentos** usando **Gemini**.
Las mismas operaciones se exponen como **MCP** para usarlas desde **Claude**.

## Arquitectura
```
core/            Tools y acceso al Excel (única fuente de verdad)
mcp/             Servidor MCP -> usar desde Claude escribiéndole
app/             Web (FastAPI + Gemini) -> audio/texto/foto/PDF
generar_finanzas.py / validar.py   Generan y validan la planilla
armar_app.py     Scaffolder (regenera core/mcp/app)
scripts/modificar_proyecto.py      CLI de mantenimiento
```

## Tools disponibles
`agregar_cobro` · `agregar_movimiento` · `actualizar_ipc` · `listar_inquilinos` · `resumen_mensual`

## Arrancar la app
```
.\run.ps1
```
http://localhost:8000 · Ver `SETUP.md`.

## Usar desde Claude (MCP)
Ver `mcp/README_MCP.md`.
'''

FILES["SETUP.md"] = r'''# Puesta en marcha

## Claves
- **GEMINI_API_KEY** (app web): sacala gratis en https://aistudio.google.com/apikey
- No hace falta nada más: Gemini procesa audio, imágenes y PDF de forma nativa.

Copiá `.env.example` a `.env` y completá `GEMINI_API_KEY`.

## App web
```
.\run.ps1
```
Abrí http://localhost:8000

## Servidor MCP (para Claude)
```
pip install -r mcp/requirements.txt
```
Configuralo en Claude Desktop (ver `mcp/README_MCP.md`).

## Notas
- Cerrá la planilla en Excel antes de procesar (Windows la bloquea al escribir).
- Cada escritura hace backup automático en `backups/`.
- Audio del navegador = webm; si Gemini lo rechaza, subí un mp3/wav.
- App pública en internet o MCP en claude.ai (nube) = requieren hosting; pedímelo.
'''

count = 0
for rel, content in FILES.items():
    p = BASE / rel
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content, encoding="utf-8")
    count += 1
print(f"OK: {count} archivos generados.")
for rel in FILES:
    print("  -", rel)
