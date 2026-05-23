"""Genera el libro de finanzas familiares 2026 (bimonetario ARS/USD) con
reajuste de alquileres por inflacion INDEC, frecuencia configurable por inquilino."""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import random

AZUL = "1F3864"; AZUL_CLARO = "D9E1F2"; GRIS = "F2F2F2"; AMARILLO = "FFF2CC"
INPUT = Font(color="0000FF"); LINK = Font(color="008000"); BOLD = Font(bold=True, color="000000")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, color=AZUL, size=16); SUB = Font(bold=True, color=AZUL, size=12)
H_FILL = PatternFill("solid", fgColor=AZUL); BAND = PatternFill("solid", fgColor=AZUL_CLARO)
GREY = PatternFill("solid", fgColor=GRIS); YEL = PatternFill("solid", fgColor=AMARILLO)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F_ARS = '"$"#,##0;("$"#,##0);"-"'; F_USD = '"US$"#,##0;("US$"#,##0);"-"'
F_PCT = '0.0%;(0.0%);"-"'; F_NUM = '#,##0.0;(#,##0.0);"-"'
F_DATE = 'DD/MM/YYYY'; F_MY = 'MMM-YYYY'; F_COEF = '0.000'; F_IDX = '#,##0.00'
MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

wb = Workbook()

def header_row(ws, row, headers, start_col=1, height=30):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER; c.border = BORDER

def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = TITLE
    if sub:
        ws["A2"] = sub; ws["A2"].font = Font(italic=True, color="595959", size=10)

def setw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def dn(name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))

# ---- crear hojas en orden ----
wsP = wb.active; wsP.title = "Portada"
wpar = wb.create_sheet("Parametros")
winf = wb.create_sheet("Inflacion INDEC")
wgui = wb.create_sheet("Guia INDEC")
wi = wb.create_sheet("Inquilinos")
wc = wb.create_sheet("Cobros")
wt = wb.create_sheet("Tablero Inquilinos")
wm = wb.create_sheet("Movimientos")
wr = wb.create_sheet("Resumen Mensual")
wb_ = wb.create_sheet("Presupuesto")
wk = wb.create_sheet("Indicadores")

# ============================================================
# PORTADA
# ============================================================
ws = wsP; ws.sheet_view.showGridLines = False
setw(ws, {"A": 3, "B": 42, "C": 60})
ws["B2"] = "FINANZAS DE LA FAMILIA"; ws["B2"].font = Font(bold=True, color=AZUL, size=22)
ws["B3"] = "Plan económico doméstico · Año 2026"; ws["B3"].font = Font(italic=True, color="595959", size=12)
ws["B5"] = "¿Cómo está organizado este libro?"; ws["B5"].font = SUB
guia = [
    ("Parámetros", "Empezá acá: tipo de cambio, mes de cálculo y tabla de frecuencias."),
    ("Inflacion INDEC", "Cargá el índice IPC cada mes. Reajusta todos los alquileres."),
    ("Guia INDEC", "Paso a paso para traer el IPC automático con Power Query."),
    ("Inquilinos", "Los 29 inquilinos. El alquiler se reajusta solo por inflación INDEC."),
    ("Cobros", "Detalle de cada cobro de alquiler (una fila por pago)."),
    ("Tablero Inquilinos", "Tablero automático: quién pagó cada mes y % de cumplimiento."),
    ("Movimientos", "Ingresos y egresos de la familia (Yo, Hermana, Mamá)."),
    ("Resumen Mensual", "Resumen automático mes a mes: ingresos, egresos y ahorro."),
    ("Presupuesto", "Presupuesto por categoría vs gasto real del año."),
    ("Indicadores", "Indicadores financieros para tomar mejores decisiones."),
]
r = 6
for name, desc in guia:
    ws.cell(row=r, column=2, value=name).font = BOLD
    ws.cell(row=r, column=3, value=desc).alignment = LEFT
    r += 1
ws.cell(row=r+1, column=2, value="Código de colores").font = SUB
leyenda = [("Azul", "Dato que cargás vos (editable)", "0000FF", None),
           ("Negro", "Cálculo / fórmula (no tocar)", "000000", None),
           ("Verde", "Trae datos de otra hoja", "008000", None),
           ("Fondo amarillo", "Supuesto clave a revisar", "000000", AMARILLO)]
r += 2
for txt, desc, color, bg in leyenda:
    cc = ws.cell(row=r, column=2, value=txt); cc.font = Font(bold=True, color=color)
    if bg: cc.fill = PatternFill("solid", fgColor=bg)
    ws.cell(row=r, column=3, value=desc).alignment = LEFT
    r += 1
ws.cell(row=r+1, column=2, value="Reajuste de alquileres:").font = BOLD
ws.cell(row=r+1, column=3, value="Cada inquilino tiene un 'mes base' y una frecuencia. El alquiler sube según el IPC del INDEC que cargás en la hoja 'Inflacion INDEC'.").alignment = LEFT

# ============================================================
# PARÁMETROS
# ============================================================
wp = wpar; wp.sheet_view.showGridLines = False
setw(wp, {"A": 32, "B": 16, "C": 3, "D": 16, "E": 14, "F": 14, "G": 18, "H": 16, "I": 26, "J": 22, "K": 3, "L": 16, "M": 10})
title(wp, "PARÁMETROS Y SUPUESTOS", "Actualizá el tipo de cambio y el mes de cálculo. Las listas alimentan los menús desplegables.")
def par(row, label, val, fmt):
    wp.cell(row=row, column=1, value=label).font = BOLD
    c = wp.cell(row=row, column=2, value=val); c.font = INPUT; c.fill = YEL; c.number_format = fmt; c.border = BORDER
par(4, "Tipo de cambio USD → ARS", 1150, F_ARS)
par(5, "Año de trabajo", 2026, '0')
par(6, "Fondo de emergencia objetivo (meses)", 6, '0')
par(7, "Saldo disponible hoy (ARS)", 0, F_ARS)
par(8, "Mes de cálculo del reajuste", date(2026, 5, 1), F_MY)

listas = {
    "D": ("Monedas", ["ARS", "USD"]),
    "E": ("Tipo mov.", ["Ingreso", "Egreso"]),
    "F": ("Miembros", ["Yo", "Hermana", "Mamá", "Familia"]),
    "G": ("Medios de pago", ["Efectivo", "Transferencia", "Cheque", "Débito", "Mercado Pago", "Otro"]),
    "H": ("Estado cobro", ["Cobrado", "Pendiente", "Parcial", "Vencido"]),
    "I": ("Categorías egreso", ["Vivienda/Expensas", "Servicios (luz/gas/agua)", "Internet/Teléfono",
            "Supermercado/Comida", "Salud/Obra social", "Transporte/Auto", "Educación",
            "Impuestos (ABL/AFIP/Rentas)", "Mantenimiento locales", "Seguros", "Honorarios/Gestión",
            "Ocio/Viajes", "Indumentaria", "Ahorro/Inversión", "Otros"]),
    "J": ("Categorías ingreso", ["Alquileres", "Sueldo", "Honorarios", "Intereses/Inversiones",
            "Venta de activos", "Otros ingresos"]),
}
hdr = 11
for col, (name, vals) in listas.items():
    c = wp[f"{col}{hdr}"]; c.value = name; c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER; c.border = BORDER
    for i, v in enumerate(vals):
        cc = wp.cell(row=hdr + 1 + i, column=ord(col) - 64, value=v); cc.font = INPUT; cc.border = BORDER; cc.alignment = LEFT
# tabla de frecuencias (label + meses)
wp.cell(row=hdr, column=12, value="Frecuencia").font = H_FONT
wp.cell(row=hdr, column=12).fill = H_FILL; wp.cell(row=hdr, column=12).alignment = CENTER; wp.cell(row=hdr, column=12).border = BORDER
wp.cell(row=hdr, column=13, value="Meses").font = H_FONT
wp.cell(row=hdr, column=13).fill = H_FILL; wp.cell(row=hdr, column=13).alignment = CENTER; wp.cell(row=hdr, column=13).border = BORDER
frecs = [("Mensual", 1), ("Trimestral", 3), ("Semestral", 6), ("Anual", 12), ("Sin ajuste", 0)]
for i, (lab, mm) in enumerate(frecs):
    wp.cell(row=hdr+1+i, column=12, value=lab).font = INPUT; wp.cell(row=hdr+1+i, column=12).border = BORDER
    wp.cell(row=hdr+1+i, column=13, value=mm).font = INPUT; wp.cell(row=hdr+1+i, column=13).border = BORDER

dn("TC", "Parametros!$B$4"); dn("ANIO", "Parametros!$B$5")
dn("FondoObj", "Parametros!$B$6"); dn("SaldoHoy", "Parametros!$B$7"); dn("MESCALC", "Parametros!$B$8")
def lista_ref(col, n): return f"Parametros!${col}${hdr+1}:${col}${hdr+n}"
dn("Monedas", lista_ref("D", 2)); dn("Tipos", lista_ref("E", 2)); dn("Miembros", lista_ref("F", 4))
dn("Medios", lista_ref("G", 6)); dn("EstadosCobro", lista_ref("H", 4))
dn("CatEgreso", lista_ref("I", 15)); dn("CatIngreso", lista_ref("J", 6))
dn("Frecuencias", f"Parametros!$L${hdr+1}:$L${hdr+5}")
dn("TablaFrec", f"Parametros!$L${hdr+1}:$M${hdr+5}")
N_CAT_E = 15; CAT_E_START = hdr + 1

# ============================================================
# INFLACION INDEC
# ============================================================
wf = winf; wf.sheet_view.showGridLines = False
title(wf, "ÍNDICE DE PRECIOS AL CONSUMIDOR — INDEC (IPC Nacional, Nivel General)",
      "Cargá en la columna 'IPC índice' el nivel que publica el INDEC cada mes. Fuente: indec.gob.ar (comunicado IPC) o datos.gob.ar (ver hoja 'Guia INDEC').")
HRF = 4
header_row(wf, HRF, ["Clave", "Mes", "IPC índice (INDEC)", "Var. mensual", "Var. acum. (vs inicio)"])
setw(wf, {"A": 10, "B": 14, "C": 20, "D": 14, "E": 18})
wf.freeze_panes = "A5"
months = []
y, mo = 2022, 1
for _ in range(60):
    months.append(date(y, mo, 1)); mo += 1
    if mo > 12: mo = 1; y += 1
first_row = HRF + 1
for i, d in enumerate(months):
    r = first_row + i
    wf.cell(row=r, column=1, value=f"=YEAR(B{r})*100+MONTH(B{r})").alignment = CENTER
    cb = wf.cell(row=r, column=2, value=d); cb.number_format = F_MY; cb.alignment = CENTER
    cc = wf.cell(row=r, column=3); cc.font = INPUT; cc.number_format = F_IDX  # IPC indice (input)
    cd = wf.cell(row=r, column=4, value=f'=IF(OR($C{r}="",$C{r-1}="",$C{r-1}=0),"",$C{r}/$C{r-1}-1)') if i > 0 else wf.cell(row=r, column=4, value="")
    cd.number_format = F_PCT
    ce = wf.cell(row=r, column=5, value=f'=IF(OR($C{r}="",$C${first_row}="",$C${first_row}=0),"",$C{r}/$C${first_row}-1)'); ce.number_format = F_PCT
    for col in range(1, 6):
        wf.cell(row=r, column=col).border = BORDER
    if i % 2 == 1:
        for col in range(1, 6):
            wf.cell(row=r, column=col).fill = GREY
LAST_F = HRF + 60
dn("ClaveIPC", f"'Inflacion INDEC'!$A${first_row}:$A${LAST_F}")
dn("IndiceIPC", f"'Inflacion INDEC'!$C${first_row}:$C${LAST_F}")

# ============================================================
# GUIA INDEC
# ============================================================
wg = wgui; wg.sheet_view.showGridLines = False
setw(wg, {"A": 3, "B": 100})
title(wg, "CÓMO ACTUALIZAR EL IPC DEL INDEC")
lines = [
    ("SUB", "Opción A — Manual (rápido, recomendado para empezar)"),
    ("", "1) Una vez por mes (INDEC publica a mediados de mes) entrá a: indec.gob.ar → Informes técnicos → Índice de precios al consumidor (IPC)."),
    ("", "2) Buscá el 'Nivel general' del IPC Nacional (es un número índice, base dic-2016=100)."),
    ("", "3) En la hoja 'Inflacion INDEC', escribí ese número en la columna 'IPC índice' del mes correspondiente."),
    ("", "4) Listo: todos los alquileres en pesos se reajustan solos según su frecuencia."),
    ("", ""),
    ("SUB", "Opción B — Automático con Power Query (API oficial datos.gob.ar)"),
    ("", "1) En Excel: pestaña Datos → Obtener datos → Desde otras fuentes → Desde la Web."),
    ("", "2) Pegá esta URL (reemplazá ID_SERIE por el id de la serie del IPC Nivel General):"),
    ("MONO", "   https://apis.datos.gob.ar/series/api/series/?ids=ID_SERIE&format=csv&start_date=2022-01"),
    ("", "3) Para conseguir el ID_SERIE: entrá a https://datos.gob.ar/series y buscá 'IPC nivel general nacional índice'. Abrí la serie y copiá el identificador (id) que figura en su ficha/URL."),
    ("", "4) Power Query te muestra una tabla con 'indice_tiempo' (fecha) y el valor del índice. Quitá las columnas que no uses y dejá fecha + valor."),
    ("", "5) Cerrar y cargar. Después, copiá los valores a la columna 'IPC índice' de la hoja 'Inflacion INDEC' (o cargá la consulta directamente sobre esa hoja)."),
    ("", "6) Para refrescar cada mes: Datos → Actualizar todo. Trae el último IPC publicado."),
    ("", ""),
    ("SUB", "Cómo se calcula el reajuste"),
    ("", "Alquiler ajustado = Alquiler base × (IPC del mes de ajuste ÷ IPC del mes base)."),
    ("", "Ejemplo: alquiler base $300.000 fijado cuando el IPC valía 100. Si hoy el IPC vale 130, el coeficiente es 1,30 y el alquiler pasa a $390.000 (+30%)."),
    ("", "La 'frecuencia' define cada cuánto salta: Mensual ajusta todos los meses; Trimestral/Semestral/Anual mantienen el valor fijo y saltan al cumplir el período, acumulando toda la inflación de ese tramo."),
    ("", "Los contratos en USD NO se ajustan por IPC (el dólar es su cobertura): se convierten a pesos con el tipo de cambio de 'Parámetros'."),
]
rr = 4
for style, txt in lines:
    c = wg.cell(row=rr, column=2, value=txt)
    if style == "SUB": c.font = SUB
    elif style == "MONO": c.font = Font(name="Consolas", color="000000")
    c.alignment = LEFT
    rr += 1

# ============================================================
# INQUILINOS  (con reajuste por inflacion)
# ============================================================
wi.sheet_view.showGridLines = False
title(wi, "LISTADO DE INQUILINOS — con reajuste automático por inflación INDEC",
      "Cargá 'Mes base', 'Alquiler base', moneda y frecuencia. El 'Alquiler ajustado (ARS)' se calcula solo con el IPC y el tipo de cambio.")
cols_i = ["ID", "Inquilino", "Local / Propiedad", "Rubro", "Inicio contrato", "Fin contrato",
          "Mes base (ajuste)", "Alquiler base", "Moneda", "Frecuencia ajuste",
          "IPC base", "Mes últ. ajuste", "IPC ajuste", "Coef.", "Alquiler ajustado",
          "Alquiler ajustado (ARS)", "Día venc.", "Depósito", "Estado", "Contacto", "Notas"]
HR = 4
header_row(wi, HR, cols_i)
setw(wi, {"A": 5, "B": 20, "C": 18, "D": 14, "E": 13, "F": 13, "G": 13, "H": 13, "I": 8,
          "J": 14, "K": 10, "L": 13, "M": 10, "N": 8, "O": 16, "P": 18, "Q": 8, "R": 12,
          "S": 11, "T": 16, "U": 22})
wi.freeze_panes = "C5"
random.seed(7)
rubros = ["Comercio", "Gastronomía", "Indumentaria", "Oficina", "Depósito", "Kiosco", "Servicios"]
frec_cycle = ["Mensual", "Trimestral", "Semestral", "Anual"]
for i in range(29):
    r = HR + 1 + i; idn = i + 1
    wi.cell(row=r, column=1, value=idn).font = INPUT
    wi.cell(row=r, column=2, value=f"Inquilino {idn:02d}").font = INPUT
    wi.cell(row=r, column=3, value=f"Local {idn:02d}").font = INPUT
    wi.cell(row=r, column=4, value=rubros[i % len(rubros)]).font = INPUT
    wi.cell(row=r, column=5, value=date(2024, (i % 12) + 1, 1)).number_format = F_DATE
    wi.cell(row=r, column=6).number_format = F_DATE
    wi.cell(row=r, column=7, value=date(2025, (i % 12) + 1, 1)).number_format = F_MY  # mes base
    wi.cell(row=r, column=7).font = INPUT
    moneda = "USD" if i % 5 == 0 else "ARS"
    base = round(random.randint(180, 650) * 1000, -3) if moneda == "ARS" else random.randint(300, 1200)
    wi.cell(row=r, column=8, value=base).font = INPUT; wi.cell(row=r, column=8).number_format = '#,##0'
    wi.cell(row=r, column=9, value=moneda).font = INPUT; wi.cell(row=r, column=9).alignment = CENTER
    wi.cell(row=r, column=10, value=frec_cycle[i % 4]).font = INPUT; wi.cell(row=r, column=10).alignment = CENTER
    wi.cell(row=r, column=11, value=f'=IFERROR(INDEX(IndiceIPC,MATCH(YEAR($G{r})*100+MONTH($G{r}),ClaveIPC,0)),0)').number_format = F_IDX
    wi.cell(row=r, column=12, value=(
        f'=IF($G{r}="","",IF(IFERROR(VLOOKUP($J{r},TablaFrec,2,0),0)=0,$G{r},'
        f'EDATE($G{r},INT(MAX(0,(YEAR(MESCALC)-YEAR($G{r}))*12+(MONTH(MESCALC)-MONTH($G{r})))/'
        f'IFERROR(VLOOKUP($J{r},TablaFrec,2,0),0))*IFERROR(VLOOKUP($J{r},TablaFrec,2,0),0))))')
    ).number_format = F_MY
    wi.cell(row=r, column=13, value=f'=IFERROR(INDEX(IndiceIPC,MATCH(YEAR($L{r})*100+MONTH($L{r}),ClaveIPC,0)),0)').number_format = F_IDX
    wi.cell(row=r, column=14, value=f'=IF(OR($K{r}=0,$M{r}=0),1,$M{r}/$K{r})').number_format = F_COEF
    wi.cell(row=r, column=15, value=f'=IF($H{r}="","",IF($I{r}="USD",$H{r},$H{r}*$N{r}))').number_format = '#,##0'
    wi.cell(row=r, column=16, value=f'=IF($O{r}="","",IF($I{r}="USD",$O{r}*TC,$O{r}))').number_format = F_ARS
    wi.cell(row=r, column=17, value=random.choice([5, 10, 1])).font = INPUT
    wi.cell(row=r, column=18).number_format = '#,##0'
    wi.cell(row=r, column=19, value="Activo").font = INPUT
    wi.cell(row=r, column=20).font = INPUT
    wi.cell(row=r, column=21).font = INPUT
    for col in range(1, 22):
        wi.cell(row=r, column=col).border = BORDER
    if i % 2 == 1:
        for col in range(1, 22):
            wi.cell(row=r, column=col).fill = GREY
LAST_I = HR + 29
tr = LAST_I + 1
wi.cell(row=tr, column=2, value="TOTAL alquiler mensual ajustado (ARS)").font = BOLD
tcc = wi.cell(row=tr, column=16, value=f"=SUM(P{HR+1}:P{LAST_I})"); tcc.font = BOLD; tcc.number_format = F_ARS
for col in range(1, 22):
    wi.cell(row=tr, column=col).fill = BAND; wi.cell(row=tr, column=col).border = BORDER
dv = DataValidation(type="list", formula1="=Monedas", allow_blank=True); wi.add_data_validation(dv); dv.add(f"I{HR+1}:I{LAST_I}")
dv2 = DataValidation(type="list", formula1="=Frecuencias", allow_blank=True); wi.add_data_validation(dv2); dv2.add(f"J{HR+1}:J{LAST_I}")
dv3 = DataValidation(type="list", formula1='"Activo,Inactivo,En mora,Vacante"', allow_blank=True); wi.add_data_validation(dv3); dv3.add(f"S{HR+1}:S{LAST_I}")

# ============================================================
# COBROS
# ============================================================
wc.sheet_view.showGridLines = False
title(wc, "REGISTRO DE COBROS DE ALQUILER", "Una fila por pago recibido. Elegí el ID del inquilino; el nombre y las conversiones se completan solos.")
cols_c = ["Fecha cobro", "ID Inq.", "Inquilino", "Mes", "Período (texto)", "Monto", "Moneda",
          "Monto (ARS)", "Monto (USD)", "Medio de pago", "Estado", "Notas"]
HRC = 4
header_row(wc, HRC, cols_c)
setw(wc, {"A": 14, "B": 8, "C": 22, "D": 7, "E": 16, "F": 14, "G": 9, "I": 14, "H": 16, "J": 16, "K": 12, "L": 26})
wc.freeze_panes = "A5"
NROWS_C = 400
for i in range(NROWS_C):
    r = HRC + 1 + i
    wc.cell(row=r, column=1).number_format = F_DATE; wc.cell(row=r, column=1).font = INPUT
    wc.cell(row=r, column=2).font = INPUT
    wc.cell(row=r, column=3, value=f'=IF($B{r}="","",IFERROR(VLOOKUP($B{r},Inquilinos!$A:$B,2,0),"ID inexistente"))').font = LINK
    wc.cell(row=r, column=4, value=f'=IF($A{r}="","",MONTH($A{r}))').alignment = CENTER
    wc.cell(row=r, column=5).font = INPUT
    wc.cell(row=r, column=6).font = INPUT; wc.cell(row=r, column=6).number_format = '#,##0'
    wc.cell(row=r, column=7).font = INPUT; wc.cell(row=r, column=7).alignment = CENTER
    wc.cell(row=r, column=8, value=f'=IF($F{r}="","",IF($G{r}="USD",$F{r}*TC,$F{r}))').number_format = F_ARS
    wc.cell(row=r, column=9, value=f'=IF($F{r}="","",IF($G{r}="USD",$F{r},$F{r}/TC))').number_format = F_USD
    wc.cell(row=r, column=10).font = INPUT; wc.cell(row=r, column=11).font = INPUT; wc.cell(row=r, column=12).font = INPUT
    for col in range(1, 13):
        wc.cell(row=r, column=col).border = BORDER
LAST_C = HRC + NROWS_C
trc = LAST_C + 1
wc.cell(row=trc, column=5, value="TOTAL cobrado").font = BOLD
wc.cell(row=trc, column=8, value=f"=SUM(H{HRC+1}:H{LAST_C})").font = BOLD; wc.cell(row=trc, column=8).number_format = F_ARS
wc.cell(row=trc, column=9, value=f"=SUM(I{HRC+1}:I{LAST_C})").font = BOLD; wc.cell(row=trc, column=9).number_format = F_USD
for col in range(1, 13):
    wc.cell(row=trc, column=col).fill = BAND; wc.cell(row=trc, column=col).border = BORDER
for rng, name in [(f"G{HRC+1}:G{LAST_C}", "=Monedas"), (f"J{HRC+1}:J{LAST_C}", "=Medios"), (f"K{HRC+1}:K{LAST_C}", "=EstadosCobro")]:
    d = DataValidation(type="list", formula1=name, allow_blank=True); wc.add_data_validation(d); d.add(rng)
did = DataValidation(type="whole", operator="between", formula1="1", formula2="29", allow_blank=True)
wc.add_data_validation(did); did.add(f"B{HRC+1}:B{LAST_C}")

# ============================================================
# TABLERO INQUILINOS
# ============================================================
wt.sheet_view.showGridLines = False
title(wt, "TABLERO DE COBROS POR INQUILINO (ARS)", "Se completa solo desde 'Cobros'. El 'Esperado' usa el alquiler ya reajustado por inflación.")
cols_t = ["ID", "Inquilino"] + MESES + ["Total cobrado", "Esperado mensual", "Esperado anual", "% Cumpl."]
HRT = 4
header_row(wt, HRT, cols_t)
wt.freeze_panes = "C5"
widths_t = {"A": 5, "B": 20, "O": 15, "P": 16, "Q": 15, "R": 10}
for idx in range(12):
    widths_t[get_column_letter(3 + idx)] = 8
setw(wt, widths_t)
for i in range(29):
    r = HRT + 1 + i; src = HR + 1 + i
    wt.cell(row=r, column=1, value=f"=Inquilinos!A{src}").font = LINK
    wt.cell(row=r, column=2, value=f"=Inquilinos!B{src}").font = LINK
    for m in range(12):
        wt.cell(row=r, column=3 + m,
            value=f'=SUMIFS(Cobros!$H:$H,Cobros!$B:$B,$A{r},Cobros!$D:$D,{m+1})').number_format = F_ARS
    wt.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").number_format = F_ARS
    wt.cell(row=r, column=15).font = BOLD
    wt.cell(row=r, column=16, value=f"=Inquilinos!P{src}").font = LINK
    wt.cell(row=r, column=16).number_format = F_ARS
    wt.cell(row=r, column=17, value=f"=P{r}*12").number_format = F_ARS
    wt.cell(row=r, column=18, value=f'=IF(Q{r}=0,"",O{r}/Q{r})').number_format = F_PCT
    for col in range(1, 19):
        wt.cell(row=r, column=col).border = BORDER
    if i % 2 == 1:
        for col in range(1, 19):
            wt.cell(row=r, column=col).fill = GREY
LAST_T = HRT + 29
trt = LAST_T + 1
wt.cell(row=trt, column=2, value="TOTAL").font = BOLD
for m in range(12):
    L = get_column_letter(3 + m)
    c = wt.cell(row=trt, column=3 + m, value=f"=SUM({L}{HRT+1}:{L}{LAST_T})"); c.number_format = F_ARS; c.font = BOLD
for col, L in [(15, "O"), (16, "P"), (17, "Q")]:
    c = wt.cell(row=trt, column=col, value=f"=SUM({L}{HRT+1}:{L}{LAST_T})"); c.number_format = F_ARS; c.font = BOLD
c = wt.cell(row=trt, column=18, value=f'=IF(Q{trt}=0,"",O{trt}/Q{trt})'); c.number_format = F_PCT; c.font = BOLD
for col in range(1, 19):
    wt.cell(row=trt, column=col).fill = BAND; wt.cell(row=trt, column=col).border = BORDER

# ============================================================
# MOVIMIENTOS
# ============================================================
wm.sheet_view.showGridLines = False
title(wm, "MOVIMIENTOS DE LA FAMILIA — INGRESOS Y EGRESOS",
      "Registrá entradas y salidas de plata. Los alquileres ya se registran en 'Cobros'; acá van sueldos, gastos, impuestos, etc.")
cols_m = ["Fecha", "Mes", "Tipo", "Miembro", "Categoría", "Descripción", "Monto", "Moneda", "Monto (ARS)", "Medio de pago", "Notas"]
HRM = 4
header_row(wm, HRM, cols_m)
wm.freeze_panes = "A5"
setw(wm, {"A": 13, "B": 7, "C": 11, "D": 12, "E": 22, "F": 26, "G": 14, "H": 9, "I": 16, "J": 16, "K": 22})
NROWS_M = 400
for i in range(NROWS_M):
    r = HRM + 1 + i
    wm.cell(row=r, column=1).number_format = F_DATE; wm.cell(row=r, column=1).font = INPUT
    wm.cell(row=r, column=2, value=f'=IF($A{r}="","",MONTH($A{r}))').alignment = CENTER
    wm.cell(row=r, column=3).font = INPUT; wm.cell(row=r, column=3).alignment = CENTER
    wm.cell(row=r, column=4).font = INPUT; wm.cell(row=r, column=5).font = INPUT; wm.cell(row=r, column=6).font = INPUT
    wm.cell(row=r, column=7).font = INPUT; wm.cell(row=r, column=7).number_format = '#,##0'
    wm.cell(row=r, column=8).font = INPUT; wm.cell(row=r, column=8).alignment = CENTER
    wm.cell(row=r, column=9, value=f'=IF($G{r}="","",IF($H{r}="USD",$G{r}*TC,$G{r}))').number_format = F_ARS
    wm.cell(row=r, column=10).font = INPUT; wm.cell(row=r, column=11).font = INPUT
    for col in range(1, 12):
        wm.cell(row=r, column=col).border = BORDER
LAST_M = HRM + NROWS_M
trm = LAST_M + 1
wm.cell(row=trm, column=6, value="TOTAL (neto ARS)").font = BOLD
wm.cell(row=trm, column=9, value=f'=SUMIFS(I{HRM+1}:I{LAST_M},C{HRM+1}:C{LAST_M},"Ingreso")-SUMIFS(I{HRM+1}:I{LAST_M},C{HRM+1}:C{LAST_M},"Egreso")').font = BOLD
wm.cell(row=trm, column=9).number_format = F_ARS
for col in range(1, 12):
    wm.cell(row=trm, column=col).fill = BAND; wm.cell(row=trm, column=col).border = BORDER
for rng, name in [(f"C{HRM+1}:C{LAST_M}", "=Tipos"), (f"D{HRM+1}:D{LAST_M}", "=Miembros"), (f"H{HRM+1}:H{LAST_M}", "=Monedas"), (f"J{HRM+1}:J{LAST_M}", "=Medios")]:
    d = DataValidation(type="list", formula1=name, allow_blank=True); wm.add_data_validation(d); d.add(rng)

# ============================================================
# RESUMEN MENSUAL
# ============================================================
wr.sheet_view.showGridLines = False
title(wr, "RESUMEN MENSUAL (ARS)", "Todo automático: alquileres desde 'Cobros', otros ingresos y egresos desde 'Movimientos'.")
cols_r = ["Mes", "N°", "Ingresos alquileres", "Otros ingresos", "Total ingresos", "Egresos", "Resultado neto", "Tasa de ahorro", "Acumulado"]
HRR = 4
header_row(wr, HRR, cols_r)
setw(wr, {"A": 8, "B": 5, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 13, "I": 18})
for m in range(12):
    r = HRR + 1 + m
    wr.cell(row=r, column=1, value=MESES[m]).alignment = CENTER
    wr.cell(row=r, column=2, value=m + 1).alignment = CENTER
    wr.cell(row=r, column=3, value=f"=SUMIFS(Cobros!$H:$H,Cobros!$D:$D,$B{r})").number_format = F_ARS
    wr.cell(row=r, column=4, value=f'=SUMIFS(Movimientos!$I:$I,Movimientos!$C:$C,"Ingreso",Movimientos!$B:$B,$B{r})').number_format = F_ARS
    wr.cell(row=r, column=5, value=f"=C{r}+D{r}").number_format = F_ARS; wr.cell(row=r, column=5).font = BOLD
    wr.cell(row=r, column=6, value=f'=SUMIFS(Movimientos!$I:$I,Movimientos!$C:$C,"Egreso",Movimientos!$B:$B,$B{r})').number_format = F_ARS
    wr.cell(row=r, column=7, value=f"=E{r}-F{r}").number_format = F_ARS; wr.cell(row=r, column=7).font = BOLD
    wr.cell(row=r, column=8, value=f'=IF(E{r}=0,"",G{r}/E{r})').number_format = F_PCT
    wr.cell(row=r, column=9, value=(f"=G{r}" if m == 0 else f"=I{r-1}+G{r}")).number_format = F_ARS
    for col in range(1, 10):
        wr.cell(row=r, column=col).border = BORDER
    if m % 2 == 1:
        for col in range(1, 10):
            wr.cell(row=r, column=col).fill = GREY
LAST_R = HRR + 12
trr = LAST_R + 1
wr.cell(row=trr, column=1, value="AÑO").font = BOLD
for col, L in [(3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")]:
    c = wr.cell(row=trr, column=col, value=f"=SUM({L}{HRR+1}:{L}{LAST_R})"); c.font = BOLD; c.number_format = F_ARS
wr.cell(row=trr, column=8, value=f'=IF(E{trr}=0,"",G{trr}/E{trr})').font = BOLD
wr.cell(row=trr, column=8).number_format = F_PCT
for col in range(1, 10):
    wr.cell(row=trr, column=col).fill = BAND; wr.cell(row=trr, column=col).border = BORDER
ANO_ING = f"'Resumen Mensual'!$E${trr}"; ANO_EGR = f"'Resumen Mensual'!$F${trr}"; ANO_NETO = f"'Resumen Mensual'!$G${trr}"

# ============================================================
# PRESUPUESTO
# ============================================================
wb_.sheet_view.showGridLines = False
title(wb_, "PRESUPUESTO ANUAL POR CATEGORÍA (ARS)", "Cargá el presupuesto mensual por categoría (azul). El gasto real se trae solo desde 'Movimientos'.")
cols_b = ["Categoría", "Presup. mensual", "Presup. anual", "Gasto real (año)", "Desvío", "% usado"]
HRB = 4
header_row(wb_, HRB, cols_b)
setw(wb_, {"A": 26, "B": 16, "C": 16, "D": 16, "E": 16, "F": 11})
cats = listas["I"][1]
for i, cat in enumerate(cats):
    r = HRB + 1 + i
    wb_.cell(row=r, column=1, value=f"=Parametros!I{CAT_E_START+i}").font = LINK
    wb_.cell(row=r, column=2).font = INPUT; wb_.cell(row=r, column=2).number_format = F_ARS
    wb_.cell(row=r, column=3, value=f"=B{r}*12").number_format = F_ARS
    wb_.cell(row=r, column=4, value=f'=SUMIFS(Movimientos!$I:$I,Movimientos!$C:$C,"Egreso",Movimientos!$E:$E,$A{r})').number_format = F_ARS
    wb_.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = F_ARS
    wb_.cell(row=r, column=6, value=f'=IF(C{r}=0,"",D{r}/C{r})').number_format = F_PCT
    for col in range(1, 7):
        wb_.cell(row=r, column=col).border = BORDER
    if i % 2 == 1:
        for col in range(1, 7):
            wb_.cell(row=r, column=col).fill = GREY
LAST_B = HRB + len(cats)
trb = LAST_B + 1
wb_.cell(row=trb, column=1, value="TOTAL").font = BOLD
for col, L in [(2, "B"), (3, "C"), (4, "D"), (5, "E")]:
    c = wb_.cell(row=trb, column=col, value=f"=SUM({L}{HRB+1}:{L}{LAST_B})"); c.font = BOLD; c.number_format = F_ARS
wb_.cell(row=trb, column=6, value=f'=IF(C{trb}=0,"",D{trb}/C{trb})').font = BOLD
wb_.cell(row=trb, column=6).number_format = F_PCT
for col in range(1, 7):
    wb_.cell(row=trb, column=col).fill = BAND; wb_.cell(row=trb, column=col).border = BORDER

# ============================================================
# INDICADORES
# ============================================================
wk.sheet_view.showGridLines = False
title(wk, "INDICADORES FINANCIEROS", "Panel para tomar decisiones. Todo se calcula desde las demás hojas.")
setw(wk, {"A": 3, "B": 40, "C": 18, "D": 50, "E": 3, "F": 18, "G": 16, "H": 12})
def kpi_block(ws, start_row, titulo, filas):
    ws.cell(row=start_row, column=2, value=titulo).font = SUB
    rr = start_row + 1
    header_row(ws, rr, ["Indicador", "Valor", "Lectura / comentario"], start_col=2, height=24)
    rr += 1
    for label, formula, fmt, comment in filas:
        ws.cell(row=rr, column=2, value=label).alignment = LEFT
        c = ws.cell(row=rr, column=3, value=formula); c.number_format = fmt; c.alignment = RIGHT
        ws.cell(row=rr, column=4, value=comment).alignment = LEFT
        for col in (2, 3, 4):
            ws.cell(row=rr, column=col).border = BORDER
        rr += 1
    return rr + 1
esp_anual_total = f"'Tablero Inquilinos'!$Q${LAST_T + 1}"; cobrado_anual_total = f"'Tablero Inquilinos'!$O${LAST_T + 1}"
r = 4
r = kpi_block(wk, r, "Resultado del año", [
    ("Ingresos totales (año)", f"={ANO_ING}", F_ARS, "Alquileres + otros ingresos."),
    ("Egresos totales (año)", f"={ANO_EGR}", F_ARS, "Todos los gastos cargados."),
    ("Resultado neto (año)", f"={ANO_NETO}", F_ARS, "Lo que la familia ahorró/consumió."),
    ("Tasa de ahorro anual", f'=IF({ANO_ING}=0,"",{ANO_NETO}/{ANO_ING})', F_PCT, "Sano > 20%. Cuanto más alto, mejor."),
    ("Ratio egresos / ingresos", f'=IF({ANO_ING}=0,"",{ANO_EGR}/{ANO_ING})', F_PCT, "Debajo de 80% deja margen de ahorro."),
])
r = kpi_block(wk, r, "Promedios mensuales", [
    ("Ingreso promedio mensual", f"=AVERAGE('Resumen Mensual'!E{HRR+1}:E{LAST_R})", F_ARS, "Promedio de los 12 meses."),
    ("Egreso promedio mensual", f"=AVERAGE('Resumen Mensual'!F{HRR+1}:F{LAST_R})", F_ARS, "Promedio de los 12 meses."),
    ("Ahorro promedio mensual", f"=AVERAGE('Resumen Mensual'!G{HRR+1}:G{LAST_R})", F_ARS, "Resultado neto medio."),
])
r = kpi_block(wk, r, "Alquileres / cobranza", [
    ("Alquiler esperado (año)", f"={esp_anual_total}", F_ARS, "Ya reajustado por inflación INDEC."),
    ("Alquiler cobrado (año)", f"={cobrado_anual_total}", F_ARS, "Total efectivamente cobrado."),
    ("Tasa de cobranza", f'=IF({esp_anual_total}=0,"",{cobrado_anual_total}/{esp_anual_total})', F_PCT, "Objetivo: cerca del 100%."),
    ("Morosidad", f'=IF({esp_anual_total}=0,"",1-{cobrado_anual_total}/{esp_anual_total})', F_PCT, "Lo que falta cobrar. Cuanto más bajo, mejor."),
    ("Ingreso medio por local (mensual)", f'=IF(COUNT(Inquilinos!P{HR+1}:P{LAST_I})=0,"",AVERAGE(Inquilinos!P{HR+1}:P{LAST_I}))', F_ARS, "Renta media reajustada por inquilino."),
    ("Concentración: mayor inquilino", f'=IF({esp_anual_total}=0,"",MAX(\'Tablero Inquilinos\'!Q{HRT+1}:Q{LAST_T})/{esp_anual_total})', F_PCT, "Si es alto, dependés mucho de un local."),
])
r = kpi_block(wk, r, "Liquidez y respaldo", [
    ("Saldo disponible hoy", "=SaldoHoy", F_ARS, "Cargalo en 'Parámetros'."),
    ("Fondo de emergencia (meses cubiertos)", f'=IF(AVERAGE(\'Resumen Mensual\'!F{HRR+1}:F{LAST_R})=0,"",SaldoHoy/AVERAGE(\'Resumen Mensual\'!F{HRR+1}:F{LAST_R}))', F_NUM, "Meses que aguantás sin ingresos."),
    ("Objetivo fondo de emergencia (ARS)", f"=FondoObj*AVERAGE('Resumen Mensual'!F{HRR+1}:F{LAST_R})", F_ARS, "Meses objetivo × egreso medio."),
    ("Brecha vs objetivo de fondo", f"=FondoObj*AVERAGE('Resumen Mensual'!F{HRR+1}:F{LAST_R})-SaldoHoy", F_ARS, "Lo que falta ahorrar para el colchón."),
])
col0 = 6; r2 = 4
wk.cell(row=r2, column=col0, value="Egresos por miembro (año)").font = SUB
header_row(wk, r2 + 1, ["Miembro", "Egresos", "% del total"], start_col=col0, height=24)
miembros = listas["F"][1]; rr = r2 + 2
for mb in miembros:
    wk.cell(row=rr, column=col0, value=mb).alignment = LEFT
    wk.cell(row=rr, column=col0+1, value=f'=SUMIFS(Movimientos!$I:$I,Movimientos!$C:$C,"Egreso",Movimientos!$D:$D,"{mb}")').number_format = F_ARS
    wk.cell(row=rr, column=col0+2, value=f'=IF({ANO_EGR}=0,"",G{rr}/{ANO_EGR})').number_format = F_PCT
    for col in (col0, col0+1, col0+2):
        wk.cell(row=rr, column=col).border = BORDER
    rr += 1
wk.cell(row=rr, column=col0, value="TOTAL").font = BOLD
wk.cell(row=rr, column=col0+1, value=f"=SUM(G{r2+2}:G{rr-1})").number_format = F_ARS
wk.cell(row=rr, column=col0+1).font = BOLD
for col in (col0, col0+1, col0+2):
    wk.cell(row=rr, column=col).fill = BAND; wk.cell(row=rr, column=col).border = BORDER
rr += 3
wk.cell(row=rr, column=col0, value="Egresos por categoría (año)").font = SUB
header_row(wk, rr + 1, ["Categoría", "Egresos", "% del total"], start_col=col0, height=24)
base_cat = rr + 2
for i, cat in enumerate(cats):
    cr = base_cat + i
    wk.cell(row=cr, column=col0, value=f"=Presupuesto!A{HRB+1+i}").font = LINK
    wk.cell(row=cr, column=col0+1, value=f"=Presupuesto!D{HRB+1+i}").number_format = F_ARS
    wk.cell(row=cr, column=col0+2, value=f'=IF({ANO_EGR}=0,"",G{cr}/{ANO_EGR})').number_format = F_PCT
    for col in (col0, col0+1, col0+2):
        wk.cell(row=cr, column=col).border = BORDER

wb.calculation.fullCalcOnLoad = True
wb.save(r"C:\Users\marti\documents\proyectojose\Finanzas_Familia_2026.xlsx")
print("OK guardado")
