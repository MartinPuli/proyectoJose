"""Fuerza recalculo al abrir y valida estaticamente todas las formulas."""
import re
from openpyxl import load_workbook

PATH = r"C:\Users\marti\documents\proyectojose\Finanzas_Familia_2026.xlsx"
wb = load_workbook(PATH)

wb.calculation.fullCalcOnLoad = True

sheets = set(wb.sheetnames)
named = set(wb.defined_names.keys())
print("Hojas:", sheets)
print("Rangos con nombre:", named)

ref_quoted = re.compile(r"'([^']+)'!")
ref_plain = re.compile(r"(?<![A-Za-z0-9_'])([A-Za-zÁÉÍÓÚÑáéíóúñ][A-Za-z0-9_]*)!")
problemas = []
total_formulas = 0
KNOWN_FUNCS = {"IF","SUM","SUMIFS","AVERAGE","VLOOKUP","IFERROR","MONTH","MAX","MIN","COUNT","COUNTA","ROUND"}

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            total_formulas += 1
            if v.count("(") != v.count(")"):
                problemas.append(f"{ws.title}!{cell.coordinate}: parentesis -> {v}")
            if "#REF" in v or "#NAME" in v:
                problemas.append(f"{ws.title}!{cell.coordinate}: error literal -> {v}")
            for m in ref_quoted.findall(v):
                if m not in sheets:
                    problemas.append(f"{ws.title}!{cell.coordinate}: hoja inexistente '{m}'")
            for m in ref_plain.findall(v):
                if m not in sheets and m not in named and m.upper() not in KNOWN_FUNCS:
                    problemas.append(f"{ws.title}!{cell.coordinate}: ref desconocida '{m}' -> {v}")

print(f"\nFormulas analizadas: {total_formulas}")
if problemas:
    print(f"PROBLEMAS ({len(problemas)}):")
    for p in problemas[:50]:
        print(" -", p)
else:
    print("OK: sin problemas estructurales en las formulas.")

wb.save(PATH)
print("Guardado con fullCalcOnLoad = True")
